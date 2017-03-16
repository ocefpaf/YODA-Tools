import os
import openpyxl
from openpyxl.cell.cell import get_column_letter, column_index_from_string
from DataLoader.Model.Abstract import iInputs
from odm2api.ODM2.models import *


class ExcelInput(iInputs):
    def __init__(self, input_file, output_file=None):
        super(ExcelInput, self).__init__()
        self.input_file = input_file

        if output_file is None:
            output_file = "export.csv"

        self.output_file = output_file
        self.workbook = None
        self.sheets = None
        self.name_ranges = None

    def parse(self, file_path=None):

        if file_path is not None:
            self.input_file = file_path

        if not os.path.isfile(self.input_file):
            print "File does not exist"
            return

        self.workbook = openpyxl.load_workbook(self.input_file, read_only=True)
        self.name_ranges = self.workbook.get_named_ranges()
        self.sheets = self.workbook.get_sheet_names()
        # print self.workbook.get_sheet_by_name(self.sheets[0])

        self.__extract_method()

    def __extract_method(self):

        if 'Methods' not in self.sheets:
            return

        method_sheet = self.workbook.get_sheet_by_name('Methods')

        # Find 'Method Information'
        row = 1
        found = False
        while not found and row < method_sheet.max_row:
            cell = method_sheet.cell(row=row, column=1)
            if cell.value is not None and 'Method Information' in cell.value:
                found = True
            row += 1

        # Find the last column that has the data
        col = 1
        while col < method_sheet.max_column:
            cell = method_sheet.cell(row=row, column=col)
            if not cell.value:
                col -= 1
                break
            col += 1

        top_left_coordinate = 'A' + str(row + 1)
        bottom_right_coordinate = get_column_letter(col) + str(method_sheet.max_row)
        method_information = method_sheet[top_left_coordinate: bottom_right_coordinate]

        return method_information

    def verify(self):
        pass

    def sendODM2Session(self):
        pass