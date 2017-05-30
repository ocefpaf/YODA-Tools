import os
import openpyxl
from odm2api.ODM2.models import *
import time
import string

import xlrd
import pandas as pd


class ExcelTimeseries():

    # https://automatetheboringstuff.com/chapter12/
    def __init__(self, input_file, **kwargs):

        self.input_file = input_file
        self.gauge = None
        self.total_rows_to_read = 0
        self.rows_read = 0

        if 'gauge' in kwargs:
            self.gauge = kwargs['gauge']

        self.workbook = None
        self.sheets = []
        self.name_ranges = None
        self.tables = {}
        self._init_data(input_file)

    def get_table_name_ranges(self):
        """
        Returns a list of the name range that have a table.
        The name range should contain the cells locations of the data.
        :rtype: list
        """
        CONST_NAME = "_Table"
        table_name_range = {}
        for name_range in self.name_ranges:
            if CONST_NAME in name_range.name:
                sheet = name_range.attr_text.split('!')[0]
                sheet = sheet.replace('\'', '')

                if sheet in table_name_range:
                    table_name_range[sheet].append(name_range)
                else:
                    table_name_range[sheet] = [name_range]

        return table_name_range

    def get_range_address(self, named_range):
        return named_range.attr_text.split('!')[1].replace('$', '')

    def _init_data(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path, read_only=True)
        self.name_ranges = self.workbook.get_named_ranges()
        self.sheets = self.workbook.get_sheet_names()


    def parse(self, session):

        self._session = session

        self.tables = self.get_table_name_ranges()

        self.parse_affiliations()
        self.parse_methods()
        self.parse_variables()
        self.parse_units()
        self.parse_processing_level()
        self.parse_sampling_feature()

        # self.parse_specimens()
        # self.parse_analysis_results()
        self.parse_data_values()


    #
    # def parse_units(self):
    #
    #     CONST_UNITS = 'Units'
    #
    #     if CONST_UNITS not in self.tables:
    #         return []
    #
    #     sheet = self.workbook.get_sheet_by_name(CONST_UNITS)
    #     tables = self.tables[CONST_UNITS]
    #
    #     units = []
    #     for table in tables:
    #         cells = sheet[table.attr_text.split('!')[1].replace('$', '')]
    #         cells = cells[1:]  # Remove the column names
    #
    #         for row in cells:
    #             unit = Units()
    #             unit.UnitsTypeCV = row[0].value
    #             unit.UnitsAbbreviation = row[1].value
    #             unit.UnitsName = row[2].value
    #             unit.UnitsLink = row[3].value
    #             units.append(unit)
    #
    #     return units
    #
    #
    # def parse_processing_level(self):
    #     CONST_PROC_LEVEL = 'Processing Levels'
    #
    #     if CONST_PROC_LEVEL not in self.tables:
    #         return []
    #
    #     sheet = self.workbook.get_sheet_by_name(CONST_PROC_LEVEL)
    #     tables = self.tables[CONST_PROC_LEVEL]
    #
    #     processing_levels = []
    #     for table in tables:
    #         cells = sheet[table.attr_text.split('!')[1].replace('$', '')]
    #         cells = cells[1:]  # Remove the column names
    #
    #         for row in cells:
    #             proc_lvl = ProcessingLevels()
    #             proc_lvl.ProcessingLevelCode = row[0].value
    #             proc_lvl.Definition = row[1].value
    #             proc_lvl.Explanation = row[2].value
    #             processing_levels.append(proc_lvl)
    #
    #     return processing_levels
    #
    # def parse_sampling_feature(self):
    #     CONST_SAMP_FEAT = 'Sampling Features'
    #
    #     if CONST_SAMP_FEAT not in self.tables:
    #         return []
    #
    #     sheet = self.workbook.get_sheet_by_name(CONST_SAMP_FEAT)
    #     tables = self.tables[CONST_SAMP_FEAT]
    #
    #     sampling_features = []
    #     for table in tables:
    #         cells = sheet[table.attr_text.split('!')[1].replace('$', '')]
    #         cells = cells[1:]  # Remove the column names
    #
    #         for row in cells:
    #             sf = SamplingFeatures()
    #             sampling_features.append(sf)
    #
    #     return sampling_features
    #
    # def parse_specimens(self):
    #     CONST_SPECIMENS = 'Specimens'
    #
    #     if CONST_SPECIMENS not in self.tables:
    #         return []
    #
    #     sheet = self.workbook.get_sheet_by_name(CONST_SPECIMENS)
    #     tables = self.tables[CONST_SPECIMENS]
    #
    #     specimens = []
    #     for table in tables:
    #         cells = sheet[table.attr_text.split('!')[1].replace('$', '')]
    #         cells = cells[1:]  # Remove the column names
    #
    #         for row in cells:
    #             sp = Specimens()
    #             sp.SamplingFeatureUUID = row[0].value
    #             sp.SamplingFeatureCode = row[1].value
    #             sp.SamplingFeatureName = row[2].value
    #             sp.SamplingFeatureDescription = row[3].value
    #             sp.SamplingFeatureTypeCV = row[4].value
    #             sp.SpecimenMediumCV = row[5].value
    #             sp.IsFieldSpecimen = row[6].value
    #             specimens.append(sp)
    #
    #     return specimens
    #
    # def parse_methods(self):
    #     # tables = self.get_tables_in_sheet('Methods')
    #
    #     CONST_METHODS = "Methods"
    #
    #     if CONST_METHODS not in self.tables:
    #         return []
    #
    #     sheet = self.workbook.get_sheet_by_name(CONST_METHODS)
    #     tables = self.tables[CONST_METHODS]
    #
    #     methods = []
    #     for table in tables:
    #         cells = sheet[table.attr_text.split('!')[1].replace('$', '')]
    #         cells = cells[1:]  # Remove the column names
    #
    #         for row in cells:
    #             method = Methods()
    #             method.MethodTypeCV = row[0].value
    #             method.MethodCode = row[1].value
    #             method.MethodName = row[2].value
    #             method.MethodDescription = row[3].value
    #             method.MethodLink = row[4].value
    #
    #             org = Organizations()
    #             org.OrganizationName = row[5].value
    #             method.OrganizationObj = org
    #
    #             methods.append(method)
    #
    #     return methods
    #
    # def get_tables_in_sheet(self, sheet_name):
    #     """
    #     :param sheet_name:
    #     :rtype: list
    #     :return:
    #     """
    #
    #     if sheet_name not in self.sheets:
    #         print "%s not in excel sheet" % sheet_name
    #         return IndexError
    #
    #     sheet = self.workbook.get_sheet_by_name(sheet_name)
    #     tables = []
    #     # sheet['A1:B3']
    #     for table in sheet._tables:
    #         top_left_cell, bottom_right_cell = table.ref.split(':')
    #         tables.append(sheet[top_left_cell: bottom_right_cell])
    #     return tables
    #
    # def parse_variables(self):
    #
    #     CONST_VARIABLES = "Variables"
    #
    #     if CONST_VARIABLES not in self.tables:
    #         return []
    #
    #     sheet = self.workbook.get_sheet_by_name(CONST_VARIABLES)
    #     tables = self.tables[CONST_VARIABLES]
    #
    #     variables = []
    #     for table in tables:
    #         cells = sheet[table.attr_text.split('!')[1].replace('$', '')]
    #         cells = cells[1:]  # Remove the column names
    #         for row in cells:
    #             var = Variables()
    #             var.VariableTypeCV = row[0].value
    #             var.VariableCode = row[1].value
    #             var.VariableNameCV = row[2].value
    #             var.VariableDefinition = row[3].value
    #             var.SpeciationCV = row[4].value
    #             var.NoDataValue = row[5].value
    #             variables.append(var)
    #
    #     return variables
    #
    # def __extract_method(self):
    #
    #     if 'Methods' not in self.sheets:
    #         return
    #
    #     method_sheet = self.workbook.get_sheet_by_name('Methods')
    #
    #     # Find 'Method Information'
    #     row = 1
    #     found = False
    #     while not found and row < method_sheet.max_row:
    #         cell = method_sheet.cell(row=row, column=1)
    #         if cell.value is not None and 'Method Information' in cell.value:
    #             found = True
    #         row += 1
    #
    #     # Find the last column that has the data
    #     col = 1
    #     while col < method_sheet.max_column:
    #         cell = method_sheet.cell(row=row, column=col)
    #         if not cell.value:
    #             col -= 1
    #             break
    #         col += 1
    #
    #     top_left_coordinate = 'A' + str(row + 1)
    #     bottom_right_coordinate = get_column_letter(col) + str(method_sheet.max_row)
    #     method_information = method_sheet[top_left_coordinate: bottom_right_coordinate]
    #
    #     return method_information
    #
    #


#
# import os
# import openpyxl
# from odm2api.ODM2.models import *
# from yodatools.converter.Abstract import iInputs
# import pandas
# import time
# import string
#
#
# class ExcelSpecimen():
#     def __init__(self, input_file, **kwargs):
#
#         self.input_file = input_file
#
#         self.gauge = None
#         self.total_rows_to_read = 0
#         self.rows_read = 0
#
#         if 'gauge' in kwargs:
#             self.gauge = kwargs['gauge']
#
#         self.workbook = None
#         self.sheets = []
#         self.name_ranges = None
#         self.tables = {}
#         self._init_data(input_file)
#
#     def get_table_name_ranges(self):
#         """
#         Returns a list of the name range that have a table.
#         The name range should contain the cells locations of the data.
#         :rtype: list
#         """
#         CONST_NAME = "_Table"
#         table_name_range = {}
#         for name_range in self.name_ranges:
#             if CONST_NAME in name_range.name:
#                 sheet, dimensions = name_range.attr_text.split('!')
#                 sheet = sheet.replace('\'', '')
#
#                 if sheet in table_name_range:
#                     table_name_range[sheet].append(name_range)
#                 else:
#                     table_name_range[sheet] = [name_range]
#
#                 self.count_number_of_rows_to_parse(dimensions=dimensions)
#
#         return table_name_range
#
#     def _init_data(self, file_path):
#         self.workbook = openpyxl.load_workbook(file_path, read_only=True)
#         self.name_ranges = self.workbook.get_named_ranges()
#         self.sheets = self.workbook.get_sheet_names()
#
    def count_number_of_rows_to_parse(self, dimensions):
        # http://stackoverflow.com/questions/1450897/python-removing-characters-except-digits-from-string
        top, bottom = dimensions.replace('$', '').split(':')
        all = string.maketrans('', '')
        nodigs = all.translate(all, string.digits)
        top = int(top.translate(all, nodigs))
        bottom = int(bottom.translate(all, nodigs))
        self.total_rows_to_read += (bottom - top)

    def get_range_address(self, named_range):
        if named_range is not None:
            return named_range.attr_text.split('!')[1].replace('$', '')
        return None
    # # def parse(self, file_path=None):
    # def parse(self, session):
    #     """
    #     If any of the methods return early, then check that they have the table ranges
    #     The table range should exist in the tables from get_table_name_range()
    #     :param :
    #     :return:
    #     """
    #     self._session = session
    #
    #     self.tables = self.get_table_name_ranges()
    #
    #     start = time.time()
    #
    #     self.parse_affiliations()
    #     self.parse_methods()
    #     self.parse_variables()
    #     self.parse_units()
    #     self.parse_processing_level()
    #     self.parse_sampling_feature()
    #     self.parse_sites()
    #     self.parse_specimens()
    #     self.parse_analysis_results()
    #
    #     # self._session.commit()
    #
    #     end = time.time()
    #     print(end - start)
    #
    #     return True

    def __updateGauge(self):
        # Objects are passed by reference in Python :)
        if not self.gauge:
            return  # No gauge was passed in, but that's ok :)

        self.rows_read += 1
        value = float(self.rows_read) / self.total_rows_to_read * 100.0
        self.gauge.SetValue(value)


    def parse_units(self):
        CONST_UNITS = 'Units'

        sheet, tables = self.get_sheet_and_table(CONST_UNITS)

        if not len(tables):
            print "No Units found"
            return

        units = []
        for table in tables:
            cells = sheet[self.get_range_address(table)]

            for row in cells:
                unit = Units()
                unit.UnitsTypeCV = row[0].value
                unit.UnitsAbbreviation = row[1].value
                unit.UnitsName = row[2].value
                unit.UnitsLink = row[3].value

                if unit.UnitsTypeCV is not None:
                    units.append(unit)

                self.__updateGauge()

        self._session.add_all(units)
        self._session.flush()

    def parse_affiliations(self):  # rename to Affiliations
        SHEET_NAME = 'People and Organizations'
        sheet, tables = self.get_sheet_and_table(SHEET_NAME)

        if not len(tables):
            print "No affiliations found"
            return []

        def parse_organizations(org_table, session):
            organizations = {}

            cells = sheet[self.get_range_address(org_table)]
            for row in cells:
                org = Organizations()
                org.OrganizationTypeCV = row[0].value
                org.OrganizationCode = row[1].value
                org.OrganizationName = row[2].value
                org.OrganizationDescription = row[3].value
                org.OrganizationLink = row[4].value
                session.add(org)
                organizations[org.OrganizationName] = org
                self.__updateGauge()

            return organizations

        def parse_authors(author_table):
            authors = []
            cells = sheet[self.get_range_address(author_table)]
            for row in cells:
                ppl = People()
                org = Organizations()
                aff = Affiliations()

                ppl.PersonFirstName = row[0].value
                ppl.PersonMiddleName = row[1].value
                ppl.PersonLastName = row[2].value

                org.OrganizationName = row[3].value
                aff.AffiliationStartDate = row[4].value
                # aff.AffiliationEndDate = row[6].value
                # aff.PrimaryPhone = row[7].value
                aff.PrimaryEmail = row[5].value
                aff.PrimaryAddress = row[6].value
                # aff.PersonLink = row[10].value

                aff.OrganizationObj = org
                aff.PersonObj = ppl

                authors.append(aff)
            return authors

        # Combine table and authors

        orgs = {}
        affiliations = []
        for table in tables:
            if 'People_Table' == table.name:
                affiliations = parse_authors(table)
            else:
                orgs = parse_organizations(table, self._session)

        self._session.flush()

        for aff in affiliations:
            if aff.OrganizationObj.OrganizationName in orgs:
                aff.OrganizationObj = orgs[aff.OrganizationObj.OrganizationName]

        self._session.add_all(affiliations)
        self._session.flush()

    def get_sheet_and_table(self, sheet_name):
        if sheet_name not in self.tables:
            return [], []
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        tables = self.tables[sheet_name]

        return sheet, tables

    def parse_processing_level(self):
        CONST_PROC_LEVEL = 'Processing Levels'
        sheet, tables = self.get_sheet_and_table(CONST_PROC_LEVEL)

        if not len(tables):
            print "No processing levels found"
            return []

        processing_levels = []
        for table in tables:
            cells = sheet[self.get_range_address(table)]

            for row in cells:
                if row[0].value is not None:
                    proc_lvl = ProcessingLevels()
                    proc_lvl.ProcessingLevelCode = row[0].value
                    proc_lvl.Definition = row[1].value
                    proc_lvl.Explanation = row[2].value
                    processing_levels.append(proc_lvl)

                self.__updateGauge()

        # return processing_levels
        self._session.add_all(processing_levels)
        self._session.flush()

    def parse_sampling_feature(self):
        SHEET_NAME = 'Sampling Features'
        sheet, tables = self.get_sheet_and_table(SHEET_NAME)

        # if SHEET_NAME not in self.tables:
        #     if 'Sites' in self.tables:
        #         SHEET_NAME = 'Sites'
        #     else:
        #         print "No sampling features/sites found"
        #         return []

        # sheet = self.workbook.get_sheet_by_name(SHEET_NAME)
        # tables = self.tables[SHEET_NAME]

        # sites_table = tables[0] if tables[0].name == 'Sites_Table' else None
        elevation_datum_range = self.workbook.get_named_range("ElevationDatum")
        spatial_ref_name_range = self.workbook.get_named_range("LatLonDatum")

        spatial_references = self.parse_spatial_reference()

        sites = []
        for table in tables:
            cells = sheet[self.get_range_address(table)]
        # cells = sheet[self.get_range_address(sites_table)]


        elevation_datum = sheet[self.get_range_address(elevation_datum_range)].value
        spatial_ref_name = sheet[self.get_range_address(spatial_ref_name_range)].value.encode('utf-8')
        spatial_references_obj = spatial_references[spatial_ref_name]

        for row in cells:
            if all([row[1].value, row[2].value, row[3].value]):# are all of the required elements present
                site = Sites()
                site.SamplingFeatureUUID = row[0].value
                site.SamplingFeatureTypeCV = row[1].value
                site.SamplingFeatureGeotypeCV= row[2].value
                site.SamplingFeatureCode = row[3].value
                site.SamplingFeatureName = row[4].value
                site.SamplingFeatureDescription = row[5].value
                site.FeatureGeometryWKT = row[6].value

                site.Elevation_m = row[7].value

                site.SiteTypeCV = row[10].value
                site.Latitude = row[11].value
                site.Longitude = row[12].value
                site.ElevationDatumCV = elevation_datum
                site.SpatialReferenceObj = spatial_references_obj

                sites.append(site)
                self.__updateGauge()

            self._session.add_all(sites)
            self._session.flush()



    def parse_spatial_reference(self):
        SHEET_NAME = "SpatialReferences"
        sheet, tables = self.get_sheet_and_table(SHEET_NAME)

        if not len(tables):
            return []

        spatial_references = {}
        for table in tables:
            cells = sheet[self.get_range_address(table)]
            for row in cells:
                sr = SpatialReferences()
                sr.SRSCode = row[0].value
                sr.SRSName = row[1].value
                sr.SRSDescription = row[2].value
                sr.SRSLink = row[3].value

                spatial_references[sr.SRSName] = sr

        return spatial_references

    def parse_methods(self):
        CONST_METHODS = "Methods"
        sheet, tables = self.get_sheet_and_table(CONST_METHODS)

        if not len(tables):
            print "No methods found"
            return []

        for table in tables:
            cells = sheet[self.get_range_address(table)]

            for row in cells:
                method = Methods()
                method.MethodTypeCV = row[0].value
                method.MethodCode = row[1].value
                method.MethodName = row[2].value
                method.MethodDescription = row[3].value
                method.MethodLink = row[4].value

                # If organization does not exist then it returns None
                org = self._session.query(Organizations).filter_by(OrganizationName=row[5].value).first()
                method.OrganizationObj = org

                if method.MethodCode:  # Cannot store empty/None objects
                    self._session.add(method)

                self.__updateGauge()

        self._session.flush()

    def parse_variables(self):

        CONST_VARIABLES = "Variables"

        if CONST_VARIABLES not in self.tables:
            print "No Variables found"
            return []

        sheet = self.workbook.get_sheet_by_name(CONST_VARIABLES)
        tables = self.tables[CONST_VARIABLES]

        for table in tables:
            cells = sheet[self.get_range_address(table)]
            for row in cells:
                var = Variables()
                var.VariableTypeCV = row[0].value
                var.VariableCode = row[1].value
                var.VariableNameCV = row[2].value
                var.VariableDefinition = row[3].value
                var.SpeciationCV = row[4].value

                if row[5].value is not None:
                    if row[5].value == 'NULL':
                        #TODO break somehow because not all required data is filled out
                        print "All Variables must contain a valid No Data Value!"
                        var.NoDataValue = None
                    else:
                        var.NoDataValue = row[5].value

                if var.NoDataValue is not None:  # NoDataValue cannot be None
                    self._session.add(var)

                self.__updateGauge()

        self._session.flush()

    def is_valid(self, iterable):
        for element in iterable:
            if not element.value:
                return False
        return True


    def parse_data_values(self):
        print "working on datavalues"
        CONST_COLUMNS = "Data Columns"
        if CONST_COLUMNS not in self.tables:
            print "No Variables found"
            return []

        sheet = self.workbook.get_sheet_by_name(CONST_COLUMNS)
        tables = self.tables[CONST_COLUMNS]

        data_values = pd.read_excel(io=self.input_file, sheetname='Data Values')
        start_date = data_values["LocalDateTime"][0].to_datetime()
        utc_offset = int(data_values["UTCOffset"][0])
        value_count = len(data_values.index)

        metadata = {}

        for table in tables:
            cells = sheet[self.get_range_address(table)]

            print "looping through datavalues"
            for row in cells:
                if self.is_valid(row):

                    action = Actions()
                    feat_act = FeatureActions()
                    act_by = ActionBy()
                    series_result = TimeSeriesResults()
                    # measure_result_value = TimeSeriesResultValues()
                    # related_action = RelatedActions()


                    # Action
                    method = self._session.query(Methods).filter_by(MethodCode=row[4].value).first()
                    action.MethodObj = method
                    #TODO ActionType
                    action.ActionTypeCV = "Observation"
                    action.BeginDateTime = start_date
                    action.BeginDateTimeUTCOffset = utc_offset

                    # Feature Actions
                    sampling_feature = self._session.query(SamplingFeatures)\
                        .filter_by(SamplingFeatureCode=row[3].value)\
                        .first()

                    feat_act.SamplingFeatureObj = sampling_feature
                    feat_act.ActionObj = action

                    # Action By
                    names = row[5].value.split(' ')
                    if len(names)>2:
                        last_name = names[2]
                    else:
                        last_name = names[1]
                    first_name = names[0]

                    person = self._session.query(People).filter_by(PersonLastName=last_name, PersonFirstName=first_name).first()
                    affiliations = self._session.query(Affiliations).filter_by(PersonID=person.PersonID).first()
                    act_by.AffiliationObj = affiliations
                    act_by.ActionObj = action
                    act_by.IsActionLead = True

                    # related_action.ActionObj = action
                    # related_action.RelationshipTypeCV = "Is child of"
                    # collectionAction = self._session.query(FeatureActions)\
                    #     .filter(FeatureActions.FeatureActionID == SamplingFeatures.SamplingFeatureID)\
                    #     .filter(SamplingFeatures.SamplingFeatureCode == row[1].value)\
                    #     .first()
                    #
                    # related_action.RelatedActionObj = collectionAction.ActionObj

                    # self._session.no_autoflush
                    self._session.flush()
                    print action

                    self._session.add(action)
                    self._session.flush()
                    self._session.add(feat_act)
                    self._session.add(act_by)
                    # self._session.add(related_action)
                    self._session.flush()
                    # Measurement Result (Different from Measurement Result Value) also creates a Result
                    variable = self._session.query(Variables).filter_by(VariableCode=row[7].value).first()
                    print row[7].value
                    print variable

                    units_for_result = self._session.query(Units).filter_by(UnitsName=row[8].value).first()
                    proc_level = self._session.query(ProcessingLevels).filter_by(ProcessingLevelCode=row[9].value).first()

                    units_for_agg = self._session.query(Units).filter_by(UnitsName=row[12].value).first()

                    series_result.IntendedTimeSpacing = row[11].value
                    series_result.IntendedTimeSpacingUnitsObj = units_for_agg
                    series_result.AggregationStatisticCV = row[13].value
                    series_result.ResultUUID = row[2].value
                    series_result.FeatureActionObj = feat_act
                    series_result.ResultTypeCV = row[6].value
                    series_result.VariableObj = variable
                    series_result.UnitsObj = units_for_result
                    series_result.ProcessingLevelObj = proc_level
                    #TODO
                    series_result.StatusCV = "Complete"
                    series_result.SampledMediumCV = row[11].value
                    series_result.ValueCount = value_count
                    #TODO
                    series_result.ResultDateTime = start_date

                    self._session.add(series_result)
                    self._session.flush()

                    # Timeseries Result Value Metadata

                    my_meta = {}
                    my_meta["Result"] = series_result
                    my_meta["CensorCodeCV"] = row[14].value
                    my_meta["QualityCodeCV"] = row[15].value
                    #TODO
                    my_meta["TimeAggregationInterval"] = series_result.IntendedTimeSpacing
                    my_meta["TimeAggregationIntervalUnitsObj"] = series_result.IntendedTimeSpacingUnitsObj

                    metadata[row[1]] = my_meta

                    # self._session.add(measure_result_value)
                    self._session.flush()

                    self.__updateGauge()

        print "convert from cross tab to serial"
        self.load_time_series_values(data_values, metadata)


    def load_time_series_values(self, timeSeries, meta_dict):
        """
        Loads TimeSeriesResultsValues into pandas DataFrame
        """
        try:
            column_labels = timeSeries[0]
            date_column = "LocalDateTime"
            utc_column = "UTCOffset"
            cross_tab = pd.DataFrame(timeSeries[1:], columns=column_labels)  # , index=date_column)

        except Exception as ex:
            return

        cross_tab.set_index([date_column, utc_column], inplace=True)

        serial = cross_tab.unstack(level=[date_column, utc_column])

        serial = serial.append(pd.DataFrame(columns=['ResultID', 'CensorCodeCV', 'QualityCodeCV', 'TimeAggregationInterval',
                                                     'TimeAggregationIntervalUnitsID'])) \
            .fillna(0) \
            .reset_index() \
            .rename(columns={0: 'DataValue'}) \
            .dropna()

        print serial.columns

        for k, v in meta_dict.iteritems():
            serial.ix[serial.level_0 == k, 'ResultID'] = v["Result"].ResultID
            serial.ix[serial.level_0 == k, 'CensorCodeCV'] = v["CensorCodeCV"]
            serial.ix[serial.level_0 == k, 'QualityCodeCV'] = v["QualityCodeCV"]
            serial.ix[serial.level_0 == k, 'TimeAggregationInterval'] = v["TimeAggregationInterval"]
            serial.ix[serial.level_0 == k, 'TimeAggregationIntervalUnitsID'] = v["TimeAggregationIntervalUnitsObj"].UnitsID

        del serial['level_0']

        # TODO does this fail for sqlite in memory
        # self._session.close()
        tablename = TimeSeriesResultValues.__tablename__
        serial.to_sql(name=tablename,
                      schema=TimeSeriesResultValues.__table_args__['schema'],
                      if_exists='append',
                      chunksize=1000,
                      con=self._engine,
                      index=False)
        self._session.commit()
        return serial