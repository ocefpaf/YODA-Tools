import sys

# import xlwings as xw
import openpyxl as xl


sys.path.append("..")

from yodatools.excelparser.YODAPy.ydt.dao import BaseXldao
import yodatools.excelparser.generate.timeseries.v0_3_2.timeseries_models as model
import math

wb = None
Sheet_Names = {
    "1": "Instructions",
    "2": "People and Organizations",
    "3": 'Dataset Citation',
    "4": 'Sampling Features',
    "5": "Related Features (optional)",
    "6": 'Methods',
    "7": 'Variables',
    "8": 'Processing Levels',
    "9": 'Units',
    "10": 'Data Columns',
    "11": 'Data Values',
    "12": 'YODA File',
    "13": 'SpatialReferences'
}

class TimeseriesXlDao(BaseXldao):

    def __init__(self, excelFile=None):
        # p = sys.path
        # fullname=p[0] + '/'+excelFile

        # self.wb = xw.Workbook(excelFile)
        # self.wb = xl.load_workbook(excelFile, data_only=True)
        self.wb = xl.load_workbook(excelFile, data_only=True, read_only=True)
        print self.wb

    def close(self):
        if wb is not None:
            self.wb.close()

    def printNamedRange(self):
        # print self.wb.names
        # for e in self.wb.names:
        #     print e
        nr= self.wb.get_named_ranges()
        print nr
        for e in nr:
            print e

    def getSheetNames(self):
        # return xw.Sheet.all(self.wb)
        return self.wb.get_sheet_names()

    # YODA metadata
    def get_yoda_header(self):
        sheet = self.wb.get_sheet_by_name(Sheet_Names["1"])
        # sheet = xw.Sheet(Sheet_Names["1"])
        if sheet is None:
            return None
        sheet
        # YODAVersion = xw.Range(sheet, 'YODAVersion').value
        # TemplateProfile = xw.Range(sheet, 'TemplateProfile').value
        # TemplateVersion = xw.Range(sheet, 'TemplateVersion').value
        # VocabUpdate = xw.Range(sheet, 'VocabUpdate').value
        YODAVersion = sheet[self.wb.get_named_range('YODAVersion').value.split('!')[1]].value
        TemplateProfile = sheet[self.wb.get_named_range('TemplateProfile').value.split('!')[1]].value
        TemplateVersion = sheet[self.wb.get_named_range('TemplateVersion').value.split('!')[1]].value
        VocabUpdate = sheet[self.wb.get_named_range('VocabUpdate').value.split('!')[1]].value
        yh_list = [YODAVersion, TemplateProfile, TemplateVersion, VocabUpdate]
        if any(yh_list):
            yh = model.YODAHeader(yh_list)
            return yh
        else:
            return None

    def get_people_dictionary(self):
        # sheet = xw.Sheet(Sheet_Names["2"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["2"])
        if sheet is None:
            return None
        # people = xw.Range(sheet,"People").value
        people = sheet[self.wb.get_named_range('People').value.split('!')[1]].value
        if people is None:
            return None
        peopleResults = {}
        for x in people:
            if any(x):
                if x[1] is not None: #check middle name
                    p_name = '%s %s %s' % (x[0], x[1], x[2])
                else:
                    p_name = '%s  %s' % (x[0],x[2])
                perObj = model.Person(x[0:3])
                peopleResults.update({p_name: perObj})
        return peopleResults

    def get_all_people(self):
        # sheet = xw.Sheet(Sheet_Names["2"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["2"])
        if sheet is None:
            return None
        people = xw.Range(sheet,"People").value
        if people is None:
            return None
        peopleResults = []
        for x in people:
            if any(x):
                peopleResults.append(model.Person(x[0:3]))
        return peopleResults

    def get_all_organizations(self):
        # sheet = xw.Sheet(Sheet_Names["2"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["2"])
        if sheet is None:
            return None
        # orgs = xw.Range(sheet,"Organizations").value
        # orgs = sheet[self.wb.get_named_range('Organizations').value.split('!')[1]].value


        # if orgs is None:
        #     return None

        orgResults = []
        # for x in orgs:
        for x in sheet.rows:

            # if any(x):
            if x[2] and x[0]:
                orgResults.append(model.Organization(x))
        return orgResults

    def get_all_affiliations(self):
        # sheet = xw.Sheet(Sheet_Names["2"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["2"])
        if sheet is None:
            return None
        # people = xw.Range(sheet,"People").value
        people = sheet[self.wb.get_named_range('People').value.split('!')[1]].value
        if people is None:
            return None

        aResults = []
        for x in people:
            if any(x):
                perObj = model.Person()
                perObj.PersonFirstName = x[0]
                perObj.PersonMiddleName = x[1]
                perObj.PersonLastName = x[2]
                orgObj = model.Organization()
                orgObj.OrganizationName = x[3]
                affObj = model.Affiliation()
                affObj.Organization = orgObj
                affObj.Person = perObj
                affObj.AffiliationStartDate = x[4]
                affObj.PrimaryEmail = x[5]
                affObj.PrimaryAddress = x[6]
                aResults.append(affObj)
        return aResults

    def get_affiliations_dictionary(self):
        sheet = self.wb.get_sheet_by_name(Sheet_Names["2"])# sheet = xw.Sheet(Sheet_Names["2"])

        if sheet is None:
            return None
        # people = xw.Range(sheet,"People").value
        people = sheet[self.wb.get_named_range('People').value.split('!')[1]].value
        if people is None:
            return None

        aResults = {}
        for x in people:
            if any(x):
                if x[1] is not None: #check middle name
                    p_name = '%s %s %s' % (x[0],x[1],x[2])
                else:
                    p_name = '%s %s' % (x[0],x[2])
                perObj = model.Person()
                perObj.PersonFirstName = x[0]
                perObj.PersonMiddleName = x[1]
                perObj.PersonLastName = x[2]
                orgObj = model.Organization()
                orgObj.OrganizationName = x[3]
                affObj = model.Affiliation()
                affObj.Organization = orgObj
                affObj.Person = perObj
                affObj.AffiliationStartDate = x[4]
                affObj.PrimaryEmail = x[5]
                affObj.PrimaryAddress = x[6]
                aResults.update({p_name.replace(" ",""): affObj})
        return aResults

    def get_dataset(self):
        # sheet = xw.Sheet(Sheet_Names["3"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["3"])
        if sheet is None:
            return None

        # DatasetUUID = xw.Range(sheet,"DatasetUUID").value
        DatasetUUID = sheet[self.wb.get_named_range("DatasetUUID").value.split("!")[1]]
        # DatasetType = xw.Range(sheet,"DatasetType").value
        DatasetType = sheet[self.wb.get_named_range("DatasetType").value.split("!")[1]]
        # DatasetCode = xw.Range(sheet,"DatasetCode").value
        DatasetCode = sheet[self.wb.get_named_range("DatasetCode").value.split("!")[1]]
        # DatasetTitle = xw.Range(sheet,"DatasetTitle").value
        DatasetTitle = sheet[self.wb.get_named_range("DatasetTitle").value.split("!")[1]]
        # DatasetAbstract = xw.Range(sheet,"DatasetAbstract").value
        DatasetAbstract = sheet[self.wb.get_named_range("DatasetAbstract").value.split("!")[1]]
        ds_list = [DatasetUUID,DatasetType,DatasetCode,DatasetTitle,DatasetAbstract]
        if any(ds_list):
            ds = model.Dataset(ds_list)
            return ds
        else:
            return None

    def get_citation(self):
        sheet = self.wb.get_sheet_by_name(Sheet_Names["3"])
        # sheet = xw.Sheet(Sheet_Names["3"])
        if sheet is None:
            return None

        # CitationTitle = xw.Range(sheet,"CitationTitle").value
        CitationTitle = sheet[self.wb.get_named_range('CitationTitle').value.split('!')[1]].value
        Publisher = xw.Range(sheet,"Publisher").value
        PublicationYear = int(xw.Range(sheet,"PublicationYear").value)
        CitationLink = xw.Range(sheet,"CitationLink").value
        ec_list = [CitationTitle,
                   Publisher,
                   PublicationYear,
                   CitationLink]
        if any(ec_list):
            ec = model.Citation(ec_list)
            return ec
        else:
            return None

    def get_datasetcitation(self):
        sheet = self.wb.get_sheet_by_name(Sheet_Names["3"])
        # sheet = xw.Sheet(Sheet_Names["3"])
        if sheet is None:
            return None

        DatasetCitationRelationship = xw.Range(sheet,"DatasetCitationRelationship").value
        if DatasetCitationRelationship is None:
            return None
        Dataset = self.get_dataset()
        Citation = self.get_citation()
        ec_list = [DatasetCitationRelationship,Dataset,Citation]
        if any(ec_list):
            ec = model.DataSetCitation(ec_list)
            return ec
        else:
            return None

    def get_all_authorlists(self):
        sheet = self.wb.get_sheet_by_name(Sheet_Names["3"])
        # sheet = xw.Sheet(Sheet_Names["3"])
        if sheet is None:
            return None
        authors = xw.Range(sheet,"AuthorList").value
        if authors is None:
            return None

        Citation = self.get_citation()
        Person_dict = self.get_people_dictionary()
        authorResults = []
        for x in authors:
            if any(x):
                Person = Person_dict.get(x[1])
                au_list = [int(x[0]),Citation,Person]
                authorResults.append(model.AuthorList(au_list))
        return authorResults

    def get_all_samplingfeatures(self):
        # sheet = xw.Sheet(Sheet_Names["4"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["4"])
        if sheet is None:
            return None

        ElevationDatum = xw.Range(sheet,'ElevationDatum').value
        samplingfeatures = xw.Range(sheet,'SamplingFeatures').value
        sfResults = []
        for sf in samplingfeatures:
            if any(sf):
                sf.insert(8,ElevationDatum)
                sfResults.append(model.SamplingFeature(sf[0:9]))
        return sfResults

    def get_all_sites(self):
        sheet = self.wb.get_sheet_by_name(Sheet_Names["4"])
        # sheet = xw.Sheet(Sheet_Names["4"])
        if sheet is None:
            return None

        ElevationDatum = xw.Range(sheet,'ElevationDatum').value
        LatLonDatum = xw.Range(sheet,'LatLonDatum').value
        sr_obj = model.SpatialReference()
        sr_obj.SRSName = LatLonDatum

        samplingfeatures = xw.Range(sheet,'SamplingFeatures').value
        sResults = []
        for sf in samplingfeatures:
            if any(sf):
                sf.insert(8,ElevationDatum)
                sf_obj = model.SamplingFeature(sf[0:9])
                s = [sf[9],sf[10],sf[11],sf_obj,sr_obj]
                sResults.append(model.Site(s))
        return sResults

    def get_all_spatialoffsets(self):
        # sheet = xw.Sheet(Sheet_Names["5"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["5"])
        if sheet is None:
            return None
        SpatialOffsets = xw.Range(sheet,'RelatedFeatures').value
        soResults = []
        for so in SpatialOffsets:
            if any(so):
                soResults.append(model.SpatialOffset(so[3:10]))
        return soResults

    def get_all_relatedfeatures(self):
        # sheet = xw.Sheet(Sheet_Names["5"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["5"])
        if sheet is None:
            return None
        rfResults = []
        RelatedFeatures = xw.Range(sheet,'RelatedFeatures').value
        for rf in RelatedFeatures:
            if any(rf):
                first_sfObj = model.SamplingFeature()
                first_sfObj.SamplingFeatureCode = rf[0]
                second_sfObj = model.SamplingFeature()
                second_sfObj.SamplingFeatureCode = rf[2]
                so_obj = None
                if rf[3] is not None:
                    so_obj = model.SpatialOffset(rf[3:10])
                rf_list = [rf[1],first_sfObj,second_sfObj,so_obj]
                rfResults.append(model.RelatedFeature(rf_list))
        return rfResults

    def get_all_methods(self):
        # sheet = xw.Sheet(Sheet_Names["6"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["6"])
        if sheet is None:
            return None
        Methods = xw.Range(sheet,'Methods').value
        mResults = []
        for m in Methods:
            if any(m):
                if m[5] is not None: #Organization
                    org_obj = model.Organization()
                    org_obj.OrganizationName = m[5]
                    m.insert(5,org_obj)
                    if m[4]:
                        m[4] = m[4].replace('\t','').strip()
                mResults.append(model.Method(m))
        return mResults

    def get_all_variables(self):
        # sheet = xw.Sheet(Sheet_Names["7"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["7"])
        if sheet is None:
            return None
        Variables = xw.Range(sheet,'Variables').value
        vResults = []
        for v in Variables:
            if any(v):
                vResults.append(model.Variable(v))
        return vResults

    def get_all_processinglevels(self):
        # sheet = xw.Sheet(Sheet_Names["8"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["8"])
        if sheet is None:
            return None
        plResults = []
        ProcessingLevels = xw.Range(sheet,'ProcessingLevelCodes')
        rowindex = ProcessingLevels.row
        colindex = ProcessingLevels.column
        plcodes = ProcessingLevels.value
        import collections
        if isinstance(plcodes, collections.Iterable):
            for pcode in plcodes:
                if pcode is not None:
                    pls = xw.Range(sheet, (rowindex,colindex), (rowindex,colindex+2)).value
                    rowindex += 1
            #for p in ProcessingLevels.value:
            #    print p
            #    if any(p):
            #        plResults.append(model.ProcessingLevel(p))

                    plResults.append(model.ProcessingLevel(pls))
        else:
            pls = xw.Range(sheet, (rowindex, colindex), (rowindex, colindex + 2)).value
            rowindex += 1
            # for p in ProcessingLevels.value:
            #    print p
            #    if any(p):
            #        plResults.append(model.ProcessingLevel(p))

            plResults.append(model.ProcessingLevel(pls))
        return plResults

    def get_all_units(self):
        # sheet = xw.Sheet(Sheet_Names["9"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["9"])
        if sheet is None:
            return None
        uResults = []
        Units = xw.Range(sheet,'Units').value
        for u in Units:
            if any(u):
                uResults.append(model.Unit(u))
        return uResults

    def get_all_spatialreferences(self):
        # sites
        sheet = self.wb.get_sheet_by_name(Sheet_Names["4"])
        # sheet = xw.Sheet(Sheet_Names["4"])
        if sheet is None:
            return None

        LatLonDatum = xw.Range(sheet,'LatLonDatum').value
        # SRSName list
        srsname_list = [LatLonDatum]

        # timeseries result
        # currently can't be found in the sheet

        # sheet = xw.Sheet(Sheet_Names["13"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["13"])
        if sheet is None:
            return None
        sr = xw.Range(sheet,'SpatialReferences').value
        srResults = []
        for s in sr:
            if any(s) and s[1] in srsname_list:
                srResults.append(model.SpatialReference(s))
        return srResults

    def get_all_datacolumns(self):
        # sheet = xw.Sheet(Sheet_Names["10"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["10"])
        if sheet is None:
            return None
        datacloumnResults = []
        # DataColumns = xw.Range(sheet,'DataColumns').value
        DataColumns= sheet.iter_rows(range_string = self.wb.get_named_range('DataColumnsTable').value.split('!')[1])
        for d in DataColumns:
            if any(d):
                datacloumnResults.append(model.DataColumn(d))
        if len(datacloumnResults) == 0:
            return None
        return datacloumnResults

    def get_begindate_and_utcoffset(self):
        # sheet = xw.Sheet(Sheet_Names["11"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["11"])
        if sheet is None:
            return None
        # bdate = xw.Range(sheet,(2,1)).value
        bdate = sheet.cell(row=2, column=1).value
        # utcoffset = int(xw.Range(sheet,(2,2)).value)
        utcoffset = int(sheet.cell(row=2, column=2).value)
        return bdate, utcoffset

    def get_all_actions(self):
        aResults = []
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None
        actiontypecv = "Observation"
        begindatetime, beginUTCOffset = self.get_begindate_and_utcoffset()
        mcodelist = []
        for dc in dc_list:
            mcode = getattr(dc,'MethodCode')
            if not mcode in mcodelist:
                mcodelist.append(mcode)

                mObj = model.Method()
                mObj.MethodCode = mcode
                aObj = model.Action()
                aObj.ActionTypeCV = actiontypecv
                aObj.BeginDateTime = begindatetime
                aObj.BeginDateTimeUTCOffset = beginUTCOffset
                aObj.Method = mObj
                aResults.append(aObj)
        return aResults

    def get_all_actionbys(self):
        aResults = []
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None
        actiontypecv = "Observation"
        begindatetime,beginUTCOffset = self.get_begindate_and_utcoffset()
        aff = self.get_affiliations_dictionary()
        idlist = []
        for dc in dc_list:
            mcode = getattr(dc,'MethodCode')
            collactor = getattr(dc,'DataCollector')
            id = '{0}:{1}'.format(mcode,collactor)
            if not id in idlist:
                idlist.append(id)

                mObj = model.Method()
                mObj.MethodCode = mcode
                aObj = model.Action()
                aObj.Method = mObj
                aObj.ActionTypeCV = actiontypecv
                aObj.BeginDateTime = begindatetime
                aObj.BeginDateTimeUTCOffset = beginUTCOffset
                affObj = aff.get(collactor.replace(" ",""))
                abObj = model.ActionBy()
                abObj.IsActionLead = True
                abObj.Action = aObj
                abObj.Affiliation = affObj
                aResults.append(abObj)

        return aResults

    def get_all_featureactions(self):
        faResults = []
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None
        actiontypecv = "Observation"
        begindatetime,beginUTCOffset = self.get_begindate_and_utcoffset()
        idlist = []
        for dc in dc_list:
            mcode = getattr(dc,'MethodCode')
            sfcode = getattr(dc,'SamplingFeatureCode')
            id = '{0}:{1}'.format(mcode,sfcode)
            if not id in idlist:
                idlist.append(id)

                sfObj = model.SamplingFeature()
                sfObj.SamplingFeatureCode = sfcode

                mObj = model.Method()
                mObj.MethodCode = mcode
                caObj = model.Action()
                caObj.Method = mObj
                caObj.ActionTypeCV = actiontypecv
                caObj.BeginDateTime = begindatetime
                caObj.BeginDateTimeUTCOffset = beginUTCOffset
                cfaObj = model.FeatureAction()
                cfaObj.SamplingFeature = sfObj
                cfaObj.Action = caObj
                faResults.append(cfaObj)
        return faResults

    def get_valuecount(self):
        # sheet = xw.Sheet(Sheet_Names["11"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["11"])
        if sheet is None:
            return None
        # v_count = xw.Range(sheet,'A1').table.last_cell.row - 1
        v_count =sheet.max_row-2

        return v_count

    def get_all_results(self):
        rResults = []
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None

        actiontypecv = "Observation"
        begindatetime,beginUTCOffset = self.get_begindate_and_utcoffset()
        valuecount = self.get_valuecount()
        for dc in dc_list:
            vObj = model.Variable()
            vObj.VariableCode = getattr(dc,'VariableCode')
            sfObj = model.SamplingFeature()
            sfObj.SamplingFeatureCode = getattr(dc,'SamplingFeatureCode')
            mObj = model.Method()
            mObj.MethodCode = getattr(dc,'MethodCode')
            caObj = model.Action()
            caObj.ActionTypeCV = actiontypecv
            caObj.BeginDateTime = begindatetime
            caObj.BeginDateTimeUTCOffset = beginUTCOffset
            caObj.Method = mObj
            faObj = model.FeatureAction()
            faObj.Action = caObj
            faObj.SamplingFeature = sfObj
            pObj = model.ProcessingLevel()
            p_code = getattr(dc,'ProcessingLevelCode')
            if isinstance(p_code,float):
                pObj.ProcessingLevelCode = str(int(p_code))
            else:
                pObj.ProcessingLevelCode = p_code
            uObj = model.Unit()
            uObj.UnitsName = getattr(dc,'UnitName')

            rObj = model.Result()
            rObj.ResultUUID = getattr(dc,'ResultUUID')
            rObj.ResultTypeCV = getattr(dc,'ResultTypeCV')
            rObj.Unit = uObj
            # rObj.ResultDateTime = None
            # rObj.ResultDateTimeUTCOffset = None
            rObj.StatusCV = getattr(dc, 'StatusCV')
            rObj.SampledMediumCV = getattr(dc,'SampledMediumCV')
            rObj.ValueCount = valuecount
            rObj.FeatureAction = faObj
            rObj.Variable = vObj
            rObj.ProcessingLevel = pObj
            # rObj.TaxonomicClassifier = None
            rResults.append(rObj)
        return rResults

    def get_all_timeseriesresults(self):
        trResults = []
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None

        actiontypecv = "Observation"
        begindatetime,beginUTCOffset = self.get_begindate_and_utcoffset()
        valuecount = self.get_valuecount()
        for dc in dc_list:
            vObj = model.Variable()
            vObj.VariableCode = getattr(dc,'VariableCode')
            sfObj = model.SamplingFeature()
            sfObj.SamplingFeatureCode = getattr(dc,'SamplingFeatureCode')
            mObj = model.Method()
            mObj.MethodCode = getattr(dc,'MethodCode')
            caObj = model.Action()
            caObj.ActionTypeCV = actiontypecv
            caObj.BeginDateTime = begindatetime
            caObj.BeginDateTimeUTCOffset = beginUTCOffset
            caObj.Method = mObj
            faObj = model.FeatureAction()
            faObj.Action = caObj
            faObj.SamplingFeature = sfObj
            pObj = model.ProcessingLevel()
            p_code = getattr(dc,'ProcessingLevelCode')
            if isinstance(p_code,float):
                pObj.ProcessingLevelCode = str(int(p_code))
            else:
                pObj.ProcessingLevelCode = p_code
            uObj = model.Unit()
            uObj.UnitsName = getattr(dc,'UnitName')

            rObj = model.Result()
            rObj.ResultUUID = getattr(dc,'ResultUUID')
            rObj.ResultTypeCV = getattr(dc,'ResultTypeCV')
            rObj.Unit = uObj
            # rObj.ResultDateTime = None
            # rObj.ResultDateTimeUTCOffset = None
            rObj.SampledMediumCV = getattr(dc,'SampledMediumCV')
            rObj.ValueCount = valuecount
            rObj.FeatureAction = faObj
            rObj.Variable = vObj
            rObj.ProcessingLevel = pObj
            # rObj.TaxonomicClassifier = None
            trObj = model.TimeSeriesResult()
            trObj.Result = rObj
            trObj.IntendedTimeSpacing = getattr(dc,'IntendedTimeSpacing',None)
            itsu = getattr(dc,'IntendedTimeSpacingUnit',None)
            if itsu:
                itsuuObj = model.Unit()
                itsuuObj.UnitsName = itsu
                trObj.IntendedTimeSpacingUnit = itsuuObj
            else:
                trObj.IntendedTimeSpacingUnit = None
            trObj.AggregationStatisticCV = getattr(dc,'AggregationStatisticCV')
            trResults.append(trObj)
        return trResults

    def get_datavalues_by_column(self,col):
        # sheet = xw.Sheet(Sheet_Names["11"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["11"])
        if sheet is None:
            return None
        # last_row = xw.Range(sheet,'A1').table.last_cell.row
        last_row = sheet.max_row-1
        # DataValues = xw.Range(sheet,(2,col),(last_row,col)).value
        # DataValues = sheet.iter_rows(range_string = self.wb.get_named_range('DataColumnsTable').value.split('!')[1])
        DataValues = sheet.iter_rows()
        return DataValues

    def get_all_timeseriesresultvalues(self):
        trvResults = []
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None

        actiontypecv = "Observation"
        begindatetime,beginUTCOffset = self.get_begindate_and_utcoffset()
        valuecount = self.get_valuecount()
        for dc in dc_list:
            vObj = model.Variable()
            vObj.VariableCode = getattr(dc, 'VariableCode')
            sfObj = model.SamplingFeature()
            sfObj.SamplingFeatureCode = getattr(dc, 'SamplingFeatureCode')
            mObj = model.Method()
            mObj.MethodCode = getattr(dc, 'MethodCode')
            caObj = model.Action()
            caObj.ActionTypeCV = actiontypecv
            caObj.BeginDateTime = begindatetime
            caObj.BeginDateTimeUTCOffset = beginUTCOffset
            caObj.Method = mObj
            faObj = model.FeatureAction()
            faObj.Action = caObj
            faObj.SamplingFeature = sfObj
            pObj = model.ProcessingLevel()
            p_code = getattr(dc, 'ProcessingLevelCode')
            if isinstance(p_code, float):
                pObj.ProcessingLevelCode = str(int(p_code))
            else:
                pObj.ProcessingLevelCode = p_code
            uObj = model.Unit()
            uObj.UnitsName = getattr(dc,'UnitName')

            rObj = model.Result()
            rObj.ResultUUID = getattr(dc,'ResultUUID')
            rObj.ResultTypeCV = getattr(dc,'ResultTypeCV')
            rObj.Unit = uObj
            # rObj.ResultDateTime = None
            # rObj.ResultDateTimeUTCOffset = None
            rObj.SampledMediumCV = getattr(dc,'SampledMediumCV')
            rObj.ValueCount = valuecount
            rObj.FeatureAction = faObj
            rObj.Variable = vObj
            rObj.ProcessingLevel = pObj
            # rObj.TaxonomicClassifier = None
            trObj = model.TimeSeriesResult()
            trObj.Result = rObj
            trObj.AggregationStatisticCV = getattr(dc,'AggregationStatisticCV')
            trvObj = model.TimeSeriesResultValue()
            trvObj.TimeSeriesResult = trObj
            trvObj.ColumnLabel = getattr(dc,'ColumnLabel')
            trvObj.CensorCodeCV = getattr(dc,'CensorCodeCV')
            trvObj.QualityCodeCV = getattr(dc,'QualityCodeCV')
            trvObj.TimeAggregationInterval = getattr(dc,'TimeAggregationInterval')
            uObj = model.Unit()
            uObj.UnitsName = getattr(dc,'TimeAggregationIntervalUnitCode')
            trvObj.TimeAggregationIntervalUnit = uObj
            trvResults.append(trvObj)
        return trvResults

    def get_all_datavalues(self):
        dc_list = self.get_all_datacolumns()
        if dc_list is None:
            return None
        datavalue_header = ['ValueDateTime','ValueDateTimeUTCOffset']
        for dc in dc_list:
            datavalue_header.append(getattr(dc,'ColumnLabel'))

        # sheet = xw.Sheet(Sheet_Names["11"])
        sheet = self.wb.get_sheet_by_name(Sheet_Names["11"])
        if sheet is None:
            return None
        start_row = 2
        start_col = 1
        last_col = len(datavalue_header)
        last_row = sheet.max_row
        # last_row = xw.Range(sheet,'A1').table.last_cell.row
        max_row = 20000
        num_partition = 0
        if last_row > max_row and last_col >= 70: # Apple event timed out error
            num_partition = int(math.ceil(float(last_row)/float(max_row)))
        datavalueResults = []
        if num_partition > 0:
            partition_row = 20001
            while num_partition > 0:
                # DataValues = xw.Range(sheet,(start_row,start_col),(partition_row,last_col)).value
                DataValues = sheet.iter_rows(min_row=start_row, max_row=partition_row, min_col=start_col, max_col=last_col)
                for dv in DataValues:
                    dvobj = model.DataValue(datavalue_header,dv)
                    datavalueResults.append(dvobj)
                start_row += max_row
                partition_row += max_row
                if partition_row > last_row:
                    partition_row = last_row
                num_partition -= 1
        else:
            # DataValues = xw.Range(sheet,(start_row,start_col),(last_row,last_col)).value
            DataValues = sheet.iter_rows(min_row=start_row, max_row=last_row, min_col=start_col, max_col=last_col)
            for dv in DataValues:
                dvobj = model.DataValue(datavalue_header,dv)
                datavalueResults.append(dvobj)
        return datavalueResults
